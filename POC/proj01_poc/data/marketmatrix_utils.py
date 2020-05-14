import json
import openpyxl
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from selenium import webdriver
import requests
import json

__matrix_file_name__ = "proj01_poc/data/V5.xlsx"
# __matrix_file_name__ = "V5.xlsx"

def get_social_matrix():
    """
    Get all soical link matrix
    locale column B
    Facebook Url column W
    Instagram Url column X
    Twitter Url column Y
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["Sheet1"]
    res = {}
    for index in range(5, 46):
        locale = sheet["C{}".format(index)].value
        facebook = sheet["BB{}".format(index)].value
        instagram = sheet["BA{}".format(index)].value
        twitter = sheet["AZ{}".format(index)].value
        # Youtube = sheet["AX{}".format(index)].value                                      # 去掉youtube的链接验证
        res[locale] = [link for link in [facebook, instagram, twitter] if link]

    wb.close()
    return res

def get_all_locale():
    """
    Get all locale from matrix
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["RYI"]

    res = []
    for index in range(4, 35):
        locale = sheet["B{}".format(index)].value.split()[0]
        res.append(locale)

    return res

def get_bike_matrix():
    """
    Get All bike info matrix
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["Sheet1"]
    bike_column = ["L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "w", "x", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO"]
    # bike_column = [col for c in [bike_category[key] for key in bike_category.keys()] for col in c]
    bike_code = {}
    for col in bike_column:
        code = sheet["{}3".format(col)].value.replace("\n", "")  # 获取到每一列的车名
        bike_code[col] = code

    res = {}
    for index in range(5, 46):
        locale = sheet["C{}".format(index)].value.split()[0]  # 获取 local
        res[locale] = {}
        res[locale]["bike"] = []
        res[locale]["attainability"] = []
        res[locale]["offer"] = []
        res[locale]["freedom"] = []
        for key in bike_column:
            # column = bike_category[key]
            # for col in column:
            #     cc = sheet["{}{}".format(col, index)].value
            #     if cc and not cc.startswith("N"):
            #         res[locale][key].append(bike_code[col])
            cc = sheet["{}{}".format(key, index)].value
            if cc and cc.startswith(("Y", "y")):
                res[locale]["bike"].append(bike_code[key])          # 获取到每个locale有哪些车
        cc1 = sheet["{}{}".format("AT", index)].value
        if cc1 and cc1.startswith("https"):
            res[locale]["attainability"].append(cc1)                # 获取locale的Attainability链接
        cc2 = sheet["{}{}".format("AU", index)].value
        if cc2 and cc2.startswith("https"):
            res[locale]["offer"].append(cc2)                        # 获取locale的offer链接
        cc3 = sheet["{}{}".format("AV", index)].value
        if cc3:
            res[locale]["freedom"].append(cc3.strip())                      # 获取locale的free

    return res




if __name__ == "__main__":
    print(json.dumps(get_bike_matrix()))

